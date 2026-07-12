"""
Разовый перенос рецептов из внешней книги «oinarri рецептуры» (xlsx-выгрузка Google Sheets)
в рабочую таблицу «Мыловарня: Учёт» (Materials/Recipes/RecipeItems).

Формат вкладок-источников (КОСМЕТИКА, ДЛЯ ЛИЦА, СКРАБЫ, СВЕЧИ и диффузоры) — повторяющийся блок:
  <название продукта>
  в т.ч. продукт:      <итоговый вес>
      <ингредиент>      <вес, г>   ...   [технология в колонке H на любой строке блока]
  в т.ч. упаковка:
      <тара>            <кол-во, шт>

Запуск:
    python scripts/import_recipes.py <путь_к_xlsx> <имя_вкладки> [--write]

Без --write только печатает план (что создастся) — ничего не пишет в Google Sheets.
С --write реально создаёт Materials/Recipes/RecipeItems в «Мыловарня: Учёт».
"""

import sys

import openpyxl

sys.path.insert(0, ".")

from core.config import MATERIAL_CATEGORY_PACKAGING  # noqa: E402
from core.inventory import add_material, read_materials  # noqa: E402
from core.recipes import add_recipe, add_recipe_item, read_recipes  # noqa: E402

SECTION_PRODUCT = "в т.ч. продукт:"
SECTION_PACKAGING = "в т.ч. упаковка:"

PRODUCES_BY_TAB = {
    "СКРАБЫ": "скраб",
    "КОСМЕТИКА": "косметика",
    "ДЛЯ ЛИЦА": "уход за лицом",
    "СВЕЧИ и диффузоры": "свеча/диффузор",
}

# Отдельный список пропуска на вкладку: точные дубли имён (разный/одинаковый состав —
# неоднозначно, какой актуальный) и записи с битыми данными в источнике. Согласовано с
# founder перед переносом, не додумываем сами.
SKIP_NAMES_BY_TAB = {
    "СКРАБЫ": {
        "Гидрофильный скраб для тела",  # дубль под одним именем, два разных состава
        "Сахарный скраб для губ",  # #REF! в исходнике, нет упаковки
        "Скраб солевой с углем",  # часть строк состава без названия ингредиента
    },
    "КОСМЕТИКА": {
        "Лосьон для тела и рук с маслом виноградных косточек и семян конопли 100 мл (стекло)",  # дубль x2
        "БЕССУЛЬФАТНОЕ ЖИДКОЕ МЫЛО",  # дубль в 3 регистрах/составах (11/10/8 ингредиентов)
        "Бессульфатное жидкое мыло",
    },
}


def _clean(value) -> str:
    """Схлопывает любые пробельные символы (включая переносы строк внутри названия) в один пробел."""
    return " ".join(str(value).split())


def load_nomenclature(wb) -> dict[str, dict]:
    """name -> {тип, цена_за_ед, поставщик}. Тип: 'косм'/'тара'/'свеч'."""
    ws = wb["Номенклатура"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tip, name = row[0], row[1]
        if not name:
            continue
        price_per_unit = row[7] if len(row) > 7 else None
        out[_clean(name)] = {"тип": tip, "цена_за_ед": price_per_unit}
    return out


def _find_column(header_row, *name_options: str) -> int:
    """Ищет колонку по заголовку, а не по жёсткому индексу — разные вкладки (КОСМЕТИКА/ДЛЯ ЛИЦА
    добавляют колонку INCI) сдвигают позиции относительно СКРАБЫ/СВЕЧИ."""
    cleaned = [_clean(v) if v is not None else "" for v in header_row]
    for wanted in name_options:
        for i, h in enumerate(cleaned):
            if h == wanted:
                return i
    raise ValueError(f"Не нашёл колонку {name_options!r} среди заголовков: {cleaned!r}")


def parse_tab(wb, tab_name: str, skip_names: set[str]) -> list[dict]:
    ws = wb[tab_name]
    header = [c.value for c in ws[1]]
    col_name = _find_column(header, "Наименование продукта")
    col_composition = _find_column(header, "Состав")
    col_qty = _find_column(header, "Объем на порц., г:")
    col_technology = _find_column(header, "Технология")

    blocks = []
    current = None
    section = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        b, c, qty = row[col_name], row[col_composition], row[col_qty]
        technology = row[col_technology] if col_technology < len(row) else None

        if b and c is None and b not in (SECTION_PRODUCT, SECTION_PACKAGING):
            # новая карточка продукта
            if current is not None and current["name"] not in skip_names:
                blocks.append(current)
            current = {"name": _clean(b), "ingredients": [], "packaging": [], "technology": None, "incomplete_rows": 0}
            section = None
            continue

        if current is None:
            continue

        if b == SECTION_PRODUCT:
            section = "ingredients"
            continue
        if b == SECTION_PACKAGING:
            section = "packaging"
            continue

        if c is not None and section in ("ingredients", "packaging"):
            if qty is None:
                # название компонента есть, граммовка/кол-во — нет (незаполненный черновик
                # в источнике). Не додумываем количество — весь рецепт уходит в пропуски.
                current["incomplete_rows"] += 1
                continue
            current[section].append((_clean(c), float(qty)))
            if technology and not current["technology"]:
                current["technology"] = _clean(technology)

    if current is not None and current["name"] not in skip_names:
        blocks.append(current)

    # блоки с пустыми/повреждёнными/неполными данными — не отдаём вызывающему коду молча
    clean = []
    for block in blocks:
        if any(not name for name, _ in block["ingredients"] + block["packaging"]):
            print(f"  [пропущен, битые данные] {block['name']!r}")
            continue
        if not block["ingredients"]:
            print(f"  [пропущен, нет состава] {block['name']!r}")
            continue
        if block["incomplete_rows"] > 0:
            print(f"  [пропущен, неполный состав: {block['incomplete_rows']} строк(и) без кол-ва] {block['name']!r}")
            continue
        clean.append(block)
    return clean


def build_material_plan(blocks: list[dict], nomenclature: dict, existing_names: set[str]) -> list[dict]:
    """Материалы, которых ещё нет в реальном листе Materials, но которые нужны для этих рецептов."""
    needed: dict[str, str] = {}  # name -> 'ingredients' | 'packaging' (откуда встретили первым)
    for block in blocks:
        for name, _ in block["ingredients"]:
            needed.setdefault(name, "ingredients")
        for name, _ in block["packaging"]:
            needed.setdefault(name, "packaging")

    plan = []
    for name, kind in needed.items():
        if name in existing_names:
            continue
        nom = nomenclature.get(name)
        if kind == "packaging" or (nom and nom["тип"] == "тара"):
            category, unit = MATERIAL_CATEGORY_PACKAGING, "шт"
        else:
            category, unit = "", "г"
        price = ""
        if nom and isinstance(nom["цена_за_ед"], (int, float)):
            price = nom["цена_за_ед"]
        plan.append({"name": name, "category": category, "unit": unit, "price": price, "matched_nomenclature": nom is not None})
    return plan


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    write = "--write" in sys.argv
    if len(args) < 2:
        print("Использование: python scripts/import_recipes.py <xlsx> <вкладка> [--write]")
        return
    xlsx_path, tab_name = args[0], args[1]
    skip_names = SKIP_NAMES_BY_TAB.get(tab_name, set())

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    nomenclature = load_nomenclature(wb)
    blocks = parse_tab(wb, tab_name, skip_names)

    existing_recipes = read_recipes()
    already_imported = {_clean(n) for n in existing_recipes["название"]} if not existing_recipes.empty else set()
    if already_imported:
        skipped = [b for b in blocks if b["name"] in already_imported]
        blocks = [b for b in blocks if b["name"] not in already_imported]
        for b in skipped:
            print(f"  [уже перенесён, пропускаю] {b['name']!r}")

    print(f"\nРецептов к переносу: {len(blocks)}")
    for b in blocks:
        print(f"  - {b['name']}  (ингредиентов: {len(b['ingredients'])}, тары: {len(b['packaging'])}, технология: {'есть' if b['technology'] else 'нет'})")

    materials = read_materials()
    existing_names = {_clean(n) for n in materials["название"]} if not materials.empty else set()

    material_plan = build_material_plan(blocks, nomenclature, existing_names)
    print(f"\nНовых Materials к созданию: {len(material_plan)}")
    for m in material_plan:
        tag = "" if m["matched_nomenclature"] else "  <- НЕТ в Номенклатуре, цена/категория не определены"
        print(f"  - {m['name']!r}: категория={m['category'] or '(пусто)'!r}, ед={m['unit']}, цена={m['price']}{tag}")

    if not write:
        print("\n(без --write ничего не записано; запусти с --write, чтобы применить)")
        return

    print("\nЗаписываю Materials...")
    name_to_id = {_clean(n): mid for n, mid in zip(materials["название"], materials["id"].astype(str))} if not materials.empty else {}
    for m in material_plan:
        name_to_id[m["name"]] = add_material(m["name"], m["category"], m["unit"], 0)

    print("Записываю Recipes/RecipeItems...")
    produces = PRODUCES_BY_TAB.get(tab_name, tab_name.lower())
    for b in blocks:
        recipe_id = add_recipe(b["name"], produces, 1, b["technology"] or "")
        for name, qty in b["ingredients"] + b["packaging"]:
            add_recipe_item(recipe_id, b["name"], name_to_id[name], name, qty)
        print(f"  + {b['name']} -> recipe_id={recipe_id}")

    print("\nГотово.")


if __name__ == "__main__":
    main()
