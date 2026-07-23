"""Раздел «Компоненты» — Materials/Transactions в Postgres (было: Sheets/pandas,
см. core/inventory.py). Остаток по-прежнему нигде не хранится статично — считается
на лету из Transaction по каждому material_id (см. CLAUDE.md).

Мультитенантность: каждый запрос фильтруется по user["company_id"] — см. CLAUDE.md
"Архитектурные принципы"."""

import io
from datetime import date
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.constants import (
    DEFAULT_MATERIAL_CATEGORIES,
    DEVELOPER,
    FOUNDER,
    TRANSACTION_ADJUSTMENT,
    TRANSACTION_EXPENSE,
    TRANSACTION_INCOME,
)

from app.db import get_db
from app.models import Material, Transaction
from app.schemas import (
    AdjustmentRequest,
    BatchIncomeRequest,
    ImportCommitRequest,
    MaterialAttrsUpdate,
    NewMaterialRequest,
    TransactionRequest,
)
from app.security import get_current_user, get_owned_or_404, require_roles

router = APIRouter(prefix="/api/ingredients", tags=["ingredients"], dependencies=[Depends(get_current_user)])

_SIGN_BY_TYPE = {TRANSACTION_INCOME: 1, TRANSACTION_EXPENSE: -1, TRANSACTION_ADJUSTMENT: 1}
_YELLOW_MULTIPLIER = 1.2


def _color(balance: float, min_stock: float) -> str:
    """Жёлтая зона существует только при min_stock > 0 — иначе интервал пуст."""
    if balance < min_stock:
        return "красный"
    if min_stock > 0 and balance <= min_stock * _YELLOW_MULTIPLIER:
        return "жёлтый"
    return "зелёный"


def _material_dict(material: Material, balance: float, last_movement: date | None) -> dict:
    min_stock = float(material.min_stock)
    return {
        "id": material.id,
        "название": material.name,
        "категория": material.category,
        "ед.измерения": material.unit,
        "мин.остаток": min_stock,
        "остаток": balance,
        "ниже минимума": balance < min_stock,
        "цвет": _color(balance, min_stock),
        "последнее движение": last_movement.isoformat() if last_movement else None,
        "себестоимость 1 шт": float(material.unit_cost) if material.unit_cost is not None else None,
        "минимальная партия для закупки": (
            float(material.min_purchase_batch_qty) if material.min_purchase_batch_qty is not None else None
        ),
        "себестоимость минимальной партии": (
            float(material.min_purchase_batch_cost) if material.min_purchase_batch_cost is not None else None
        ),
        "вес минимальной партии": (
            float(material.min_purchase_batch_weight) if material.min_purchase_batch_weight is not None else None
        ),
        "поставщик": material.supplier or "",
        "INCI": material.inci or "",
        "архив": material.archived,
        "тип тары": material.packaging_type,
        "ширина, мм": float(material.width_mm) if material.width_mm is not None else None,
        "высота, мм": float(material.height_mm) if material.height_mm is not None else None,
        "длина, мм": float(material.length_mm) if material.length_mm is not None else None,
        "объём, мл": float(material.volume_ml) if material.volume_ml is not None else None,
        "материал исполнения": material.material_finish or "",
        "особенность ленты": material.tape_feature or "",
    }


def _transaction_dict(tx: Transaction) -> dict:
    return {
        "id": tx.id,
        "дата": tx.date.isoformat(),
        "material_id": tx.material_id,
        "тип": tx.type,
        "кол-во": float(tx.qty),
        "цена": float(tx.price) if tx.price is not None else "",
        "транспортные расходы": float(tx.freight_cost) if tx.freight_cost is not None else "",
        "recipe_id": tx.recipe_id or "",
        "комментарий": tx.comment or "",
    }


def _balances_and_last_movement(db: Session, company_id: str) -> tuple[dict[str, float], dict[str, date]]:
    balances: dict[str, float] = {}
    last_movement: dict[str, date] = {}
    for tx in db.scalars(select(Transaction).where(Transaction.company_id == company_id)):
        sign = _SIGN_BY_TYPE.get(tx.type, 0)
        balances[tx.material_id] = balances.get(tx.material_id, 0.0) + float(tx.qty) * sign
        if tx.material_id not in last_movement or tx.date > last_movement[tx.material_id]:
            last_movement[tx.material_id] = tx.date
    return balances, last_movement


def _get_own_material(db: Session, material_id: str, company_id: str) -> Material:
    return get_owned_or_404(db, Material, material_id, company_id, "Компонент не найден.")


@router.get("")
def list_ingredients(
    archived: bool = Query(False), user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> list[dict]:
    materials = db.scalars(
        select(Material).where(Material.company_id == user["company_id"], Material.archived == archived)
    ).all()
    balances, last_movement = _balances_and_last_movement(db, user["company_id"])
    return [_material_dict(m, balances.get(m.id, 0.0), last_movement.get(m.id)) for m in materials]


@router.get("/categories")
def list_categories(user: dict = Depends(get_current_user), db: Session = Depends(get_db)) -> list[str]:
    """Базовая тройка категорий + любые кастомные, реально использующиеся в этой компании
    (см. NewIngredientModal.tsx "+ новая категория") — категория остаётся свободной строкой,
    отдельного справочника в БД нет."""
    used = {
        c for (c,) in db.execute(
            select(Material.category).where(Material.company_id == user["company_id"]).distinct()
        ).all()
        if c
    }
    return sorted(used | set(DEFAULT_MATERIAL_CATEGORIES))


@router.get("/{material_id}/transactions")
def list_transactions(
    material_id: str, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> list[dict]:
    _get_own_material(db, material_id, user["company_id"])
    rows = db.scalars(
        select(Transaction)
        .where(Transaction.material_id == material_id, Transaction.company_id == user["company_id"])
        .order_by(Transaction.date.desc())
    )
    return [_transaction_dict(tx) for tx in rows]


@router.post("")
def create_ingredient(
    body: NewMaterialRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    material = Material(
        company_id=user["company_id"], name=body.name, category=body.category, unit=body.unit, min_stock=body.min_stock,
        packaging_type=body.packaging_type, width_mm=body.width_mm, height_mm=body.height_mm,
        length_mm=body.length_mm, volume_ml=body.volume_ml, material_finish=body.material_finish,
        tape_feature=body.tape_feature,
    )
    db.add(material)
    db.flush()
    if body.initial_qty > 0:
        db.add(
            Transaction(
                company_id=user["company_id"], material_id=material.id, type=TRANSACTION_INCOME,
                qty=body.initial_qty, comment="начальный остаток",
            )
        )
    db.commit()
    return {"id": material.id}


@router.patch("/{material_id}", dependencies=[Depends(require_roles(FOUNDER, DEVELOPER))])
def update_ingredient_attrs(
    material_id: str, body: MaterialAttrsUpdate, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    """Закупочные поля карточки (себестоимость/мин.партия/поставщик/INCI) — не
    остаток и не движение, тех правят через приход/расход/корректировку."""
    material = _get_own_material(db, material_id, user["company_id"])
    for field, value in body.model_dump(exclude_unset=True).items():
        setattr(material, field, value)
    db.commit()
    return {"ok": True}


def _require_positive(qty: float) -> None:
    if qty <= 0:
        raise HTTPException(400, "Количество должно быть больше нуля.")


@router.post("/{material_id}/income")
def add_income(
    material_id: str, body: TransactionRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    _require_positive(body.qty)
    _get_own_material(db, material_id, user["company_id"])
    db.add(
        Transaction(
            company_id=user["company_id"], material_id=material_id, type=TRANSACTION_INCOME,
            qty=body.qty, price=body.price, comment=body.comment,
        )
    )
    db.commit()
    return {"ok": True}


@router.post("/income/batch")
def add_income_batch(
    body: BatchIncomeRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    """Групповой приход одной поставки — несколько материалов сразу + одна общая сумма
    транспортных расходов, которая делится между материалами пропорционально весу
    (2026-07-21, запрос Александра). Вес берётся из "мин.партия"/"вес мин.партии" на
    карточке компонента (unit_weight = вес/кол-во минимальной партии); если хотя бы
    одно из двух не заполнено — fallback на кол-во из этой поставки (пропорция по
    штукам/кг вместо реального веса, документированное упрощение, не ошибка)."""
    if not body.items:
        raise HTTPException(400, "Список материалов пуст.")

    material_ids = [i.material_id for i in body.items]
    if len(set(material_ids)) != len(material_ids):
        raise HTTPException(400, "Один материал указан в поставке дважды.")

    if body.transport_cost < 0:
        raise HTTPException(400, "Транспортные расходы не могут быть отрицательными.")

    materials: dict[str, Material] = {}
    for item in body.items:
        _require_positive(item.qty)
        materials[item.material_id] = _get_own_material(db, item.material_id, user["company_id"])

    weights: dict[str, float] = {}
    for item in body.items:
        material = materials[item.material_id]
        if material.min_purchase_batch_weight and material.min_purchase_batch_qty:
            unit_weight = float(material.min_purchase_batch_weight) / float(material.min_purchase_batch_qty)
        else:
            unit_weight = 1.0  # fallback: пропорция по кол-ву, не по весу
        weights[item.material_id] = unit_weight * item.qty
    total_weight = sum(weights.values())

    for item in body.items:
        freight_share = (weights[item.material_id] / total_weight) * body.transport_cost if total_weight > 0 else 0.0
        db.add(
            Transaction(
                company_id=user["company_id"], material_id=item.material_id, type=TRANSACTION_INCOME,
                qty=item.qty, price=item.price, freight_cost=round(freight_share, 2) if body.transport_cost else None,
                comment=body.comment,
            )
        )
    db.commit()
    return {"ok": True, "created": len(body.items)}


@router.post("/{material_id}/expense")
def add_expense(
    material_id: str, body: TransactionRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    _require_positive(body.qty)
    _get_own_material(db, material_id, user["company_id"])
    db.add(
        Transaction(
            company_id=user["company_id"], material_id=material_id, type=TRANSACTION_EXPENSE,
            qty=body.qty, comment=body.comment,
        )
    )
    db.commit()
    return {"ok": True}


@router.post("/{material_id}/adjustment")
def add_adjustment(
    material_id: str, body: AdjustmentRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    _get_own_material(db, material_id, user["company_id"])
    balances, _ = _balances_and_last_movement(db, user["company_id"])
    current = balances.get(material_id, 0.0)
    delta = body.actual_qty - current
    if delta == 0:
        return {"ok": True, "delta": 0}
    db.add(
        Transaction(
            company_id=user["company_id"], material_id=material_id, type=TRANSACTION_ADJUSTMENT,
            qty=delta, comment=body.comment or "инвентаризация",
        )
    )
    db.commit()
    return {"ok": True, "delta": delta}


@router.get("/export-template")
def export_template(user: dict = Depends(get_current_user), db: Session = Depends(get_db)) -> StreamingResponse:
    """Шаблон для массовой инвентаризации: те же названия, что видит Founder на
    экране, плюс пустая колонка «Новый остаток» — заполняется руками и грузится
    обратно через /import/preview."""
    materials = db.scalars(select(Material).where(Material.company_id == user["company_id"])).all()
    balances, _ = _balances_and_last_movement(db, user["company_id"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Компоненты"
    ws.append(["Название", "Категория", "Ед.измерения", "Текущий остаток", "Новый остаток"])
    for m in materials:
        ws.append([m.name, m.category, m.unit, balances.get(m.id, 0.0), None])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Content-Disposition — только latin-1, кириллицу передаём через filename* (RFC 5987).
    filename_utf8 = quote("компоненты.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"ingredients.xlsx\"; filename*=UTF-8''{filename_utf8}"},
    )


@router.get("/export")
def export_ingredients(user: dict = Depends(get_current_user), db: Session = Depends(get_db)) -> StreamingResponse:
    """Экспорт для просмотра/пересылки (например, в Google Таблицы) — без
    служебной колонки «Новый остаток» из /export-template."""
    materials = db.scalars(select(Material).where(Material.company_id == user["company_id"])).all()
    balances, last_movement = _balances_and_last_movement(db, user["company_id"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Компоненты"
    ws.append(["Название", "Категория", "Ед.измерения", "Остаток", "Мин.остаток", "Обновлено"])
    for m in materials:
        lm = last_movement.get(m.id)
        ws.append([m.name, m.category, m.unit, balances.get(m.id, 0.0), float(m.min_stock), lm.isoformat() if lm else None])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename_utf8 = quote("компоненты.xlsx")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"ingredients.xlsx\"; filename*=UTF-8''{filename_utf8}"},
    )


def _parse_import_file(content: bytes) -> list[dict]:
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Не удалось прочитать файл. Поддерживается формат .xlsx.")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Файл пуст.")

    header = [str(c).strip().casefold() if c is not None else "" for c in rows[0]]
    col_name = next((i for i, h in enumerate(header) if h == "название"), None)
    col_new = next((i for i, h in enumerate(header) if h == "новый остаток"), None)
    if col_new is None:
        col_new = next((i for i, h in enumerate(header) if h == "остаток"), None)
    if col_name is None or col_new is None:
        raise HTTPException(400, "В файле нет колонок «Название» и «Новый остаток» (или «Остаток»).")

    parsed = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        name = row[col_name]
        if name is None or str(name).strip() == "":
            continue
        parsed.append({"name": str(name).strip(), "new_qty": row[col_new]})
    return parsed


@router.post("/import/preview")
async def import_preview(
    file: UploadFile = File(...), user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> list[dict]:
    content = await file.read()
    parsed = _parse_import_file(content)

    materials = db.scalars(select(Material).where(Material.company_id == user["company_id"])).all()
    by_name: dict[str, list[Material]] = {}
    for m in materials:
        by_name.setdefault(m.name.strip().casefold(), []).append(m)
    balances, _ = _balances_and_last_movement(db, user["company_id"])

    result = []
    for item in parsed:
        matches = by_name.get(item["name"].casefold(), [])
        new_qty = item["new_qty"]
        row = {
            "name": item["name"],
            "material_id": None,
            "current_qty": None,
            "new_qty": None,
            "delta": None,
            "status": "",
        }
        if len(matches) == 0:
            row["status"] = "не найден"
        elif len(matches) > 1:
            row["status"] = "неоднозначное совпадение"
        elif not isinstance(new_qty, (int, float)):
            row["status"] = "не число"
        else:
            m = matches[0]
            current = balances.get(m.id, 0.0)
            row.update(
                material_id=m.id,
                current_qty=current,
                new_qty=float(new_qty),
                delta=float(new_qty) - current,
                status="ok",
            )
        result.append(row)
    return result


@router.post("/import/commit")
def import_commit(
    body: ImportCommitRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    """Строки прилетают уже подтверждённые фронтом после превью — здесь без
    повторного парсинга файла, только запись движений."""
    balances, _ = _balances_and_last_movement(db, user["company_id"])
    own_material_ids = set(
        db.scalars(select(Material.id).where(Material.company_id == user["company_id"]))
    )
    applied = 0
    for row in body.rows:
        if row.material_id not in own_material_ids:
            raise HTTPException(404, "Компонент не найден.")
        current = balances.get(row.material_id, 0.0)
        delta = row.new_qty - current
        if delta == 0:
            continue
        db.add(
            Transaction(
                company_id=user["company_id"],
                material_id=row.material_id,
                type=TRANSACTION_ADJUSTMENT,
                qty=delta,
                comment=body.comment or "импорт из файла",
            )
        )
        balances[row.material_id] = row.new_qty
        applied += 1
    db.commit()
    return {"ok": True, "applied": applied}
