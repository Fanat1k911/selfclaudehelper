"""Карточки готовых изделий — доступно только founder/developer (см. таблицу ролей в CLAUDE.md).
Обязательные поля при создании — название/категория/GTIN, см. app.constants.PRODUCT_REQUIRED_FIELDS.
Название рецепта в Product больше не хранится в БД — join при чтении.

Мультитенантность: каждый запрос фильтруется по user["company_id"]. _ready_to_ship_by_recipe
и _sold_by_product принимают company_id явно — их переиспользует sales.py."""

import io
import re
from collections import defaultdict
from datetime import date

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.constants import DEVELOPER, FOUNDER, PRODUCT_REQUIRED_FIELDS

from app.db import get_db
from app.models import Product, ProductionLog, Recipe, Sale
from app.schemas import NewProductRequest, ProductImportCommitRequest, UpdateProductRequest
from app.security import get_current_user, get_owned_or_404, require_roles

router = APIRouter(prefix="/api/products", tags=["products"], dependencies=[Depends(require_roles(FOUNDER, DEVELOPER))])


def _missing_required_fields(fields: dict[str, str]) -> list[str]:
    return [field for field in PRODUCT_REQUIRED_FIELDS if not fields.get(field, "").strip()]


def _ready_to_ship_by_recipe(db: Session, company_id: str) -> dict[str, float]:
    """Готово к отгрузке = произведено по журналу смен (кол-во продукта) минус брак,
    минус то, что уже продано (см. CLAUDE.md — пока связь рецепт-продукт всегда 1:1)."""
    produced: dict[str, float] = defaultdict(float)
    for log in db.scalars(select(ProductionLog).where(ProductionLog.company_id == company_id)):
        produced[log.recipe_id] += float(log.qty) - float(log.defects)
    return produced


def _sold_by_product(db: Session, company_id: str) -> dict[str, float]:
    sold: dict[str, float] = defaultdict(float)
    for sale in db.scalars(select(Sale).where(Sale.company_id == company_id)):
        sold[sale.product_id] += float(sale.qty)
    return sold


def _product_dict(product: Product, produced_by_recipe: dict[str, float], sold_by_product: dict[str, float]) -> dict:
    ready = None
    if product.recipe_id:
        ready = produced_by_recipe.get(product.recipe_id, 0.0) - sold_by_product.get(product.id, 0.0)
    return {
        "id": product.id,
        "название": product.name,
        "категория": product.category,
        "GTIN": product.gtin,
        "состав": product.composition or "",
        "recipe_id": product.recipe_id or "",
        "название рецепта": product.recipe.name if product.recipe_id and product.recipe else "",
        "ТН ВЭД": product.tn_ved or "",
        "декларация соответствия": product.declaration or "",
        "срок действия РД": product.declaration_expires.isoformat() if product.declaration_expires else "",
        "готово к отгрузке": ready,
    }


@router.get("")
def list_products(user: dict = Depends(get_current_user), db: Session = Depends(get_db)) -> list[dict]:
    produced_by_recipe = _ready_to_ship_by_recipe(db, user["company_id"])
    sold_by_product = _sold_by_product(db, user["company_id"])
    stmt = select(Product).where(Product.company_id == user["company_id"])
    return [_product_dict(p, produced_by_recipe, sold_by_product) for p in db.scalars(stmt)]


@router.post("")
def create_product(body: NewProductRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)) -> dict:
    fields = {
        "название": body.name,
        "категория": body.category,
        "GTIN": body.gtin,
        "состав": body.composition,
        "recipe_id": body.recipe_id,
        "ТН ВЭД": body.tn_ved,
        "декларация соответствия": body.declaration,
        "срок действия РД": body.declaration_expires,
    }
    missing = _missing_required_fields(fields)
    if missing:
        raise HTTPException(400, f"Не заполнены обязательные поля: {', '.join(missing)}.")

    if body.recipe_id:
        recipe = get_owned_or_404(db, Recipe, body.recipe_id, user["company_id"], "Рецепт не найден.")
        if recipe.archived:
            raise HTTPException(400, "Рецепт в архиве, привязать к продукту нельзя.")

    product = Product(
        company_id=user["company_id"],
        name=body.name,
        category=body.category,
        gtin=body.gtin,
        composition=body.composition or None,
        recipe_id=body.recipe_id or None,
        tn_ved=body.tn_ved or None,
        declaration=body.declaration or None,
        declaration_expires=date.fromisoformat(body.declaration_expires) if body.declaration_expires else None,
    )
    db.add(product)
    db.commit()
    return {"id": product.id}


@router.patch("/{product_id}")
def update_product(
    product_id: str, body: UpdateProductRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    """Единственный способ прикрепить recipe_id к продукту, созданному без него (ручной
    ввод без рецепта или Excel-импорт) — без этого эндпоинта «готово к отгрузке» навсегда
    остаётся невозможным для такого продукта (see CLAUDE.md, найдено на живых данных
    2026-07-19: у всех 28 импортированных продуктов recipe_id был пуст, и продажа не
    проверяла остаток вообще)."""
    product = get_owned_or_404(db, Product, product_id, user["company_id"], "Продукт не найден.")

    fields = {
        "название": body.name if body.name is not None else product.name,
        "категория": body.category if body.category is not None else product.category,
        "GTIN": body.gtin if body.gtin is not None else product.gtin,
    }
    missing = _missing_required_fields(fields)
    if missing:
        raise HTTPException(400, f"Не заполнены обязательные поля: {', '.join(missing)}.")

    if body.recipe_id:
        recipe = get_owned_or_404(db, Recipe, body.recipe_id, user["company_id"], "Рецепт не найден.")
        if recipe.archived:
            raise HTTPException(400, "Рецепт в архиве, привязать к продукту нельзя.")

    if body.name is not None:
        product.name = body.name
    if body.category is not None:
        product.category = body.category
    if body.gtin is not None:
        product.gtin = body.gtin
    if body.composition is not None:
        product.composition = body.composition or None
    if body.recipe_id is not None:
        product.recipe_id = body.recipe_id or None
    if body.tn_ved is not None:
        product.tn_ved = body.tn_ved or None
    if body.declaration is not None:
        product.declaration = body.declaration or None
    if body.declaration_expires is not None:
        product.declaration_expires = date.fromisoformat(body.declaration_expires) if body.declaration_expires else None

    db.commit()
    return {"id": product.id}


def _find_col(header: list[str], *needles: str) -> int | None:
    return next((i for i, h in enumerate(header) if any(n in h for n in needles)), None)


def _extract_expiry(text: str) -> str:
    """Из свободного текста вида «Выдан\\nDD.MM.YYYY\\nИстекает\\nDD.MM.YYYY» берём
    последнюю дату — считаем её датой окончания действия декларации."""
    dates = re.findall(r"(\d{2})\.(\d{2})\.(\d{4})", text)
    if not dates:
        return ""
    d, m, y = dates[-1]
    return f"{y}-{m}-{d}"


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _load_workbook(content: bytes):
    try:
        return load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Не удалось прочитать файл. Поддерживается формат .xlsx.")


def _parse_import_file(content: bytes, sheet_name: str | None = None) -> list[dict]:
    wb = _load_workbook(content)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(400, f"В файле нет вкладки «{sheet_name}».")
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Файл пуст.")

    header = [str(c).strip().casefold() if c is not None else "" for c in rows[0]]
    col_name = _find_col(header, "наименование", "название")
    col_category = _find_col(header, "категория")
    col_gtin = _find_col(header, "gtin", "шк")
    col_tnved = _find_col(header, "тн вэд", "тнвэд")
    col_declaration = _find_col(header, "декларац")
    col_expires = _find_col(header, "срок")

    if col_name is None or col_category is None or col_gtin is None:
        raise HTTPException(400, "В файле нет колонок «Наименование», «Категория» и «GTIN».")

    parsed = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        name = _cell_to_str(row[col_name])
        if not name:
            continue
        expiry_raw = _cell_to_str(row[col_expires]) if col_expires is not None else ""
        parsed.append(
            {
                "name": name,
                "category": _cell_to_str(row[col_category]),
                "gtin": _cell_to_str(row[col_gtin]),
                "tn_ved": _cell_to_str(row[col_tnved]) if col_tnved is not None else "",
                "declaration": _cell_to_str(row[col_declaration]) if col_declaration is not None else "",
                "declaration_expires": _extract_expiry(expiry_raw),
            }
        )
    return parsed


@router.post("/import/sheets")
async def import_sheets(file: UploadFile = File(...)) -> dict:
    content = await file.read()
    wb = _load_workbook(content)
    return {"sheets": wb.sheetnames}


@router.post("/import/preview")
async def import_preview(
    file: UploadFile = File(...), sheet_name: str | None = Form(None),
    user: dict = Depends(get_current_user), db: Session = Depends(get_db),
) -> list[dict]:
    content = await file.read()
    parsed = _parse_import_file(content, sheet_name)

    # GTIN-дедуп намеренно per-company (не глобальный, как до мультитенантности) — компании
    # не видят чужой каталог вообще, так что предупреждать про совпадение с чужим GTIN тут
    # было бы утечкой факта существования чужого товара. Разных компаний с одинаковым
    # реальным штрихкодом в этой системе не бывает на практике (разные производители).
    existing_gtins = {p.gtin for p in db.scalars(select(Product).where(Product.company_id == user["company_id"]))}
    seen_gtins: set[str] = set()

    result = []
    for item in parsed:
        row = {**item, "status": "ok"}
        missing = _missing_required_fields(
            {"название": item["name"], "категория": item["category"], "GTIN": item["gtin"]}
        )
        if missing:
            row["status"] = f"не заполнены поля: {', '.join(missing)}"
        elif item["gtin"] in existing_gtins:
            row["status"] = "GTIN уже есть в базе"
        elif item["gtin"] in seen_gtins:
            row["status"] = "дубликат GTIN в файле"
        else:
            seen_gtins.add(item["gtin"])
        result.append(row)
    return result


@router.post("/import/commit")
def import_commit(
    body: ProductImportCommitRequest, user: dict = Depends(get_current_user), db: Session = Depends(get_db)
) -> dict:
    """Строки прилетают уже подтверждённые фронтом после превью — здесь без
    повторного парсинга файла, только создание карточек продукта (импорт только
    создаёт новые продукты, обновление существующих по GTIN не поддерживается).

    recipe_id (2026-07-19) — опциональный выбор рецепта на строку с фронта (не
    авто-сопоставление по тексту из Excel — ненадёжно, могло бы молча привязать не тот
    рецепт). Без recipe_id продукт создаётся как раньше, «готово к отгрузке» для него не
    считается (см. update_product выше — можно прикрепить рецепт позже)."""
    existing_gtins = {p.gtin for p in db.scalars(select(Product).where(Product.company_id == user["company_id"]))}
    valid_recipe_ids = {
        r.id for r in db.scalars(select(Recipe).where(Recipe.company_id == user["company_id"], Recipe.archived.is_(False)))
    }
    applied = 0
    for row in body.rows:
        if row.gtin in existing_gtins:
            continue
        product = Product(
            company_id=user["company_id"],
            name=row.name,
            category=row.category,
            gtin=row.gtin,
            recipe_id=row.recipe_id if row.recipe_id in valid_recipe_ids else None,
            tn_ved=row.tn_ved or None,
            declaration=row.declaration or None,
            declaration_expires=date.fromisoformat(row.declaration_expires) if row.declaration_expires else None,
        )
        db.add(product)
        existing_gtins.add(row.gtin)
        applied += 1
    db.commit()
    return {"ok": True, "applied": applied}
